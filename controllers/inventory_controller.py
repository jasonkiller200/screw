from flask import Blueprint, jsonify, request
from models.inventory import CurrentInventory, InventoryTransaction, StockCount, get_taipei_time
from models.part import Part, Warehouse
import csv
import io
from services.inventory_service import InventoryService # Import the new service

inventory_api_bp = Blueprint('inventory_api', __name__, url_prefix='/api/inventory')

# 倉庫管理 API
@inventory_api_bp.route('/warehouses', methods=['GET'])
def get_warehouses():
    """取得所有倉庫"""
    warehouses = Warehouse.get_all()
    return jsonify(warehouses)

# 庫存查詢 API
@inventory_api_bp.route('/stock', methods=['GET'])
def get_inventory():
    """取得庫存清單"""
    warehouse_id = request.args.get('warehouse_id', type=int)
    inventories = CurrentInventory.get_all_inventory(warehouse_id)
    return jsonify(inventories)

@inventory_api_bp.route('/stock/<string:part_number>', methods=['GET'])
def get_part_stock(part_number):
    """取得特定零件的庫存"""
    warehouse_id = request.args.get('warehouse_id', type=int)
    
    # 先取得零件資訊
    part = Part.get_by_part_number(part_number)
    if not part:
        return jsonify({'error': 'Part not found'}), 404
    
    # 取得庫存資訊
    stock_info = CurrentInventory.get_current_stock(part['id'], warehouse_id)
    
    if not stock_info:
        return jsonify({'error': 'No stock information found'}), 404
    
    return jsonify({
        'part_info': part,
        'stock_info': stock_info if warehouse_id else stock_info
    })

@inventory_api_bp.route('/low-stock', methods=['GET'])
def get_low_stock():
    """取得低庫存項目"""
    warehouse_id = request.args.get('warehouse_id', type=int)
    low_stock_items = CurrentInventory.get_low_stock_items(warehouse_id)
    return jsonify(low_stock_items)

@inventory_api_bp.route('/stock/export', methods=['GET'])
def export_inventory():
    """匯出庫存清單為 CSV"""
    from flask import Response
    
    warehouse_id = request.args.get('warehouse_id', type=int)
    inventories = CurrentInventory.get_all_inventory(warehouse_id)
    
    # 生成 CSV 內容
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 寫入標題行
    writer.writerow([
        '零件編號', '零件名稱', '倉庫編號', '倉庫名稱', '現有庫存', 
        '可用庫存', '安全庫存', '補貨點', '單位', '最後更新時間'
    ])
    
    # 寫入資料行
    for inventory in inventories:
        writer.writerow([
            inventory.get('part_number', ''),
            inventory.get('part_name', ''),
            inventory.get('warehouse_code', ''),
            inventory.get('warehouse_name', ''),
            inventory.get('quantity_on_hand', 0),
            inventory.get('available_quantity', 0),
            inventory.get('safety_stock', 0),
            inventory.get('reorder_point', 0),
            inventory.get('unit', ''),
            inventory.get('last_updated', '')
        ])
    
    # 取得 CSV 內容並加上 UTF-8 BOM
    csv_content = output.getvalue()
    csv_bytes = '\ufeff' + csv_content  # UTF-8 BOM
    
    # 生成檔案名稱（使用英文避免編碼問題）
    warehouse_filter = f"_warehouse{warehouse_id}" if warehouse_id else "_all_warehouses"
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"inventory_list{warehouse_filter}_{timestamp}.csv"
    
    return Response(
        csv_bytes.encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )

@inventory_api_bp.route('/low-stock/export', methods=['GET'])
def export_low_stock():
    """匯出低庫存清單為 CSV"""
    from flask import Response
    
    warehouse_id = request.args.get('warehouse_id', type=int)
    low_stock_items = CurrentInventory.get_low_stock_items(warehouse_id)
    
    # 生成 CSV 內容
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 寫入標題行
    writer.writerow([
        '零件編號', '零件名稱', '倉庫編號', '倉庫名稱', '現有庫存', 
        '可用庫存', '安全庫存', '補貨點', '缺口數量', '單位', '建議補貨'
    ])
    
    # 寫入資料行
    for item in low_stock_items:
        shortage = max(0, item.get('reorder_point', 0) - item.get('available_quantity', 0))
        suggested_order = max(shortage, item.get('safety_stock', 0) * 2)
        
        writer.writerow([
            item.get('part_number', ''),
            item.get('part_name', ''),
            item.get('warehouse_code', ''),
            item.get('warehouse_name', ''),
            item.get('quantity_on_hand', 0),
            item.get('available_quantity', 0),
            item.get('safety_stock', 0),
            item.get('reorder_point', 0),
            shortage,
            item.get('unit', ''),
            suggested_order
        ])
    
    # 取得 CSV 內容並加上 UTF-8 BOM
    csv_content = output.getvalue()
    csv_bytes = '\ufeff' + csv_content  # UTF-8 BOM
    
    # 生成檔案名稱（使用英文避免編碼問題）
    warehouse_filter = f"_warehouse{warehouse_id}" if warehouse_id else "_all_warehouses"
    from datetime import datetime
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"low_stock_list{warehouse_filter}_{timestamp}.csv"
    
    return Response(
        csv_bytes.encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': f'attachment; filename="{filename}"',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )

# 交易記錄 API
@inventory_api_bp.route('/transactions', methods=['GET'])
def get_transactions():
    """取得交易記錄"""
    part_id = request.args.get('part_id', type=int)
    warehouse_id = request.args.get('warehouse_id', type=int)
    limit = request.args.get('limit', 100, type=int)
    
    transactions = InventoryTransaction.get_transactions(part_id, warehouse_id, limit)
    return jsonify(transactions)

@inventory_api_bp.route('/transaction-summary/<int:part_id>', methods=['GET'])
def get_transaction_summary(part_id):
    """取得交易摘要"""
    warehouse_id = request.args.get('warehouse_id', type=int)
    days = request.args.get('days', 30, type=int)
    
    summary = InventoryTransaction.get_transaction_summary(part_id, warehouse_id, days)
    return jsonify(summary)

# 入庫 API
@inventory_api_bp.route('/stock-in', methods=['POST'])
def stock_in():
    """入庫作業"""
    data = request.get_json()
    
    required_fields = ['part_number', 'warehouse_id', 'quantity', 'transaction_type']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    part_number = data['part_number']
    warehouse_id = data['warehouse_id']
    quantity = data['quantity']
    transaction_type = data['transaction_type']  # 'IN_PURCHASE', 'IN_TRANSFER', 'IN_RETURN'
    reference_type = data.get('reference_type', 'MANUAL')
    reference_id = data.get('reference_id')
    notes = data.get('notes', '')
    
    # 驗證零件是否存在
    part = Part.get_by_part_number(part_number)
    if not part:
        return jsonify({'error': 'Part not found'}), 404
    
    # 驗證數量
    try:
        quantity = int(quantity)
        if quantity <= 0:
            raise ValueError("Quantity must be positive")
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid quantity'}), 400
    
    # 驗證交易類型
    valid_in_types = ['IN_PURCHASE', 'IN_TRANSFER', 'IN_RETURN']
    if transaction_type not in valid_in_types:
        return jsonify({'error': 'Invalid transaction type for stock in'}), 400
    
    # 執行入庫
    success = CurrentInventory.update_stock(
        part.id, warehouse_id, quantity, transaction_type,
        reference_type, reference_id, notes
    )
    
    if success:
        return jsonify({
            'success': True,
            'message': f'{part_number} 入庫 {quantity} {part.unit} 成功'
        }), 201
    else:
        return jsonify({'error': 'Stock in operation failed'}), 500

# 出庫 API
@inventory_api_bp.route('/stock-out', methods=['POST'])
def stock_out():
    """出庫作業"""
    data = request.get_json()
    
    required_fields = ['part_number', 'warehouse_id', 'quantity', 'transaction_type']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'Missing required field: {field}'}), 400
    
    part_number = data['part_number']
    warehouse_id = data['warehouse_id']
    quantity = data['quantity']
    transaction_type = data['transaction_type']  # 'OUT_ISSUE', 'OUT_TRANSFER', 'OUT_SCRAP'
    reference_type = data.get('reference_type', 'MANUAL')
    reference_id = data.get('reference_id')
    notes = data.get('notes', '')
    
    try:
        # 驗證零件是否存在
        part = Part.get_by_part_number(part_number)
        if not part:
            return jsonify({'error': 'Part not found'}), 404
        
        # 驗證數量
        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError("Quantity must be positive")
        except (ValueError, TypeError):
            return jsonify({'error': 'Invalid quantity'}), 400
        
        # 驗證交易類型
        valid_out_types = ['OUT_ISSUE', 'OUT_TRANSFER', 'OUT_SCRAP']
        if transaction_type not in valid_out_types:
            return jsonify({'error': 'Invalid transaction type for stock out'}), 400
        
        # 檢查庫存是否足夠
        current_stock = CurrentInventory.get_current_stock(part.id, warehouse_id)
        if not current_stock or current_stock['available_quantity'] < quantity:
            return jsonify({
                'error': f'Insufficient stock. Available: {current_stock["available_quantity"] if current_stock else 0}'
            }), 400
        
        # 執行出庫（負數量）
        success = CurrentInventory.update_stock(
            part.id, warehouse_id, -quantity, transaction_type,
            reference_type, reference_id, notes
        )
        
        if success:
            return jsonify({
                'success': True,
                'message': f'{part_number} 出庫 {quantity} {part.unit} 成功'
            }), 201
        else:
            return jsonify({'error': 'Stock out operation failed'}), 500
    except Exception as e:
        # Log the exception for debugging purposes (in a real app, use a proper logger)
        print(f"Error during stock out: {e}")
        return jsonify({'error': f'Internal server error: {str(e)}'}), 500

# 盤點管理 API
@inventory_api_bp.route('/stock-counts', methods=['GET'])
def get_stock_counts():
    """取得所有盤點記錄"""
    counts = StockCount.get_all_counts()
    return jsonify(counts)

@inventory_api_bp.route('/stock-counts/<int:count_id>', methods=['GET'])
def get_stock_count_details(count_id):
    """取得盤點明細"""
    count_info = StockCount.get_count_by_id(count_id)
    if not count_info:
        return jsonify({'error': 'Stock count not found'}), 404
    
    details = StockCount.get_count_details(count_id)
    
    return jsonify({
        'count_info': count_info,
        'details': details
    })

@inventory_api_bp.route('/stock-counts/<int:count_id>/details/<int:detail_id>', methods=['PUT'])
def update_count_detail(count_id, detail_id):
    """更新盤點明細"""
    data = request.get_json()
    
    counted_quantity = data.get('counted_quantity')
    notes = data.get('notes', '')
    
    if counted_quantity is None:
        return jsonify({'error': 'counted_quantity is required'}), 400
    
    try:
        counted_quantity = int(counted_quantity)
        if counted_quantity < 0:
            raise ValueError("Quantity cannot be negative")
    except (ValueError, TypeError):
        return jsonify({'error': 'Invalid counted_quantity'}), 400
    
    success = StockCount.update_count_detail(detail_id, counted_quantity, notes)
    
    if success:
        return jsonify({'success': True, 'message': 'Count detail updated successfully'})
    else:
        return jsonify({'error': 'Failed to update count detail'}), 500

@inventory_api_bp.route('/stock-counts/<int:count_id>/complete', methods=['POST'])
def complete_stock_count(count_id):
    """完成盤點"""
    data = request.get_json()
    
    verified_by = data.get('verified_by', '')
    apply_adjustments = data.get('apply_adjustments', False)
    
    success = StockCount.complete_count(count_id, verified_by, apply_adjustments)
    
    if success:
        message = 'Stock count completed successfully'
        if apply_adjustments:
            message += ' and adjustments applied'
        return jsonify({'success': True, 'message': message})
    else:
        return jsonify({'error': 'Failed to complete stock count'}), 500

# 建立盤點
@inventory_api_bp.route('/stock-counts', methods=['POST'])
def create_stock_count():
    """建立新盤點"""
    data = request.get_json()
    
    warehouse_id = data.get('warehouse_id')
    count_type = data.get('count_type', 'full')
    count_date = data.get('count_date')
    counted_by = data.get('counted_by', '')
    notes = data.get('notes', '')
    
    if not warehouse_id:
        return jsonify({'error': 'Warehouse ID is required'}), 400
    
    count_id = StockCount.create_count(
        warehouse_id=warehouse_id,
        count_type=count_type,
        description=notes,
        counted_by=counted_by
    )
    
    if count_id:
        return jsonify({'success': True, 'count_id': count_id})
    else:
        return jsonify({'error': 'Failed to create stock count'}), 500

# 開始盤點
@inventory_api_bp.route('/stock-counts/<int:count_id>/start', methods=['POST'])
def start_stock_count(count_id):
    """開始盤點"""
    success = StockCount.start_count(count_id)
    
    if success:
        return jsonify({'success': True, 'message': 'Stock count started'})
    else:
        return jsonify({'error': 'Failed to start stock count'}), 500

# 刪除盤點
@inventory_api_bp.route('/stock-counts/<int:count_id>/delete', methods=['POST'])
def delete_stock_count(count_id):
    """刪除盤點（僅規劃中的盤點可刪除）"""
    from extensions import db
    from models.inventory import StockCountDetail
    
    count = StockCount.query.get(count_id)
    if not count:
        return jsonify({'error': 'Stock count not found'}), 404
    
    # 只允許刪除「規劃中」的盤點
    if count.status != 'planning':
        return jsonify({'error': 'Only planning status counts can be deleted'}), 400
    
    try:
        # 刪除相關的盤點明細
        StockCountDetail.query.filter_by(stock_count_id=count_id).delete()
        # 刪除盤點記錄
        db.session.delete(count)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Stock count deleted'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

# 盤點項目管理
@inventory_api_bp.route('/stock-counts/<int:count_id>/items', methods=['POST'])
def add_count_item(count_id):
    """新增盤點項目"""
    data = request.get_json()
    
    part_id = data.get('part_id')
    actual_quantity = data.get('actual_quantity')
    
    if not part_id or actual_quantity is None:
        return jsonify({'error': 'Part ID and actual quantity are required'}), 400
    
    success = StockCount.add_count_item(count_id, part_id, actual_quantity)
    
    if success:
        return jsonify({'success': True, 'message': 'Count item added'})
    else:
        return jsonify({'error': 'Failed to add count item'}), 500

@inventory_api_bp.route('/stock-counts/<int:count_id>/items/<int:part_id>', methods=['PUT'])
def update_count_item(count_id, part_id):
    """更新盤點項目"""
    data = request.get_json()
    
    actual_quantity = data.get('actual_quantity')
    
    if actual_quantity is None:
        return jsonify({'error': 'Actual quantity is required'}), 400
    
    success = StockCount.update_count_item(count_id, part_id, actual_quantity)
    
    if success:
        return jsonify({'success': True, 'message': 'Count item updated'})
    else:
        return jsonify({'error': 'Failed to update count item'}), 500

@inventory_api_bp.route('/stock-counts/<int:count_id>/items/update-by-part', methods=['POST'])
def update_count_item_by_part_number(count_id):
    """根據零件編號更新盤點項目"""
    from models.inventory import StockCountDetail
    from extensions import db
    
    data = request.get_json()
    part_number = data.get('part_number')
    counted_quantity = data.get('counted_quantity')
    notes = data.get('notes', '')
    
    if not part_number or counted_quantity is None:
        return jsonify({'error': 'Part number and counted quantity are required'}), 400
    
    # 查找零件
    part = Part.get_by_part_number(part_number)
    if not part:
        return jsonify({'error': f'Part not found: {part_number}'}), 404
    
    # 查找盤點明細
    detail = StockCountDetail.query.filter_by(
        stock_count_id=count_id,
        part_id=part.id
    ).first()
    
    if not detail:
        return jsonify({'error': f'Count item not found for part: {part_number}'}), 404
    
    try:
        detail.counted_quantity = counted_quantity
        detail.variance_quantity = counted_quantity - detail.system_quantity
        if notes:
            detail.notes = notes
        detail.counted_at = get_taipei_time()
        
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Count item updated'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@inventory_api_bp.route('/import-count-data', methods=['POST'])
def import_count_data_batch():
    """批量匯入盤點資料"""
    warehouse_id = request.form.get('warehouse_id')
    
    if not warehouse_id:
        return jsonify({'error': 'Warehouse ID is required'}), 400
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '' or not file.filename:
        return jsonify({'error': 'No file selected'}), 400
    
    if not file.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Only CSV files are supported'}), 400
    
    # Pass the file stream to the service layer
    result = InventoryService.import_stock_count_data(int(warehouse_id), file.stream)
    
    if result['success']:
        response_data = {
            'success': True,
            'message': f"成功匯入 {result['processed_count']} 筆盤點資料。總計 {result['total_rows']} 筆。",
            'count_id': result['count_id'],
            'processed_count': result['processed_count'],
            'total_rows': result['total_rows'],
            'errors': result['errors']
        }
        return jsonify(response_data)
    else:
        return jsonify({'success': False, 'error': result['error']}), 500

@inventory_api_bp.route('/count-template', methods=['GET'])
def download_general_count_template():
    """下載通用盤點範本"""
    from flask import Response
    
    # 生成範例 CSV 內容
    output = io.StringIO()
    writer = csv.writer(output)
    
    # 寫入標題行
    writer.writerow(['part_number', 'part_name', 'unit', 'system_quantity', 'counted_quantity', 'notes'])
    
    # 寫入範例資料
    writer.writerow(['P001', '螺絲 M6x20', '個', '100', '', '請填入實盤數量'])
    writer.writerow(['P002', '墊片 φ6', '個', '50', '', ''])
    writer.writerow(['P003', '螺帽 M6', '個', '80', '', ''])
    
    # 取得 CSV 內容並加上 UTF-8 BOM
    csv_content = output.getvalue()
    
    # 編碼為 UTF-8 並加上 BOM
    csv_bytes = '\ufeff' + csv_content  # UTF-8 BOM
    
    return Response(
        csv_bytes.encode('utf-8'),
        mimetype='text/csv; charset=utf-8',
        headers={
            'Content-Disposition': 'attachment; filename="stock_count_template.csv"',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )

@inventory_api_bp.route('/stock-counts/<int:count_id>/export', methods=['GET'])
def export_count_template(count_id):
    """匯出盤點模板（Excel 格式）"""
    from flask import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    count_info = StockCount.get_count_by_id(count_id)
    if not count_info:
        return jsonify({'error': 'Stock count not found'}), 404
    
    details = StockCount.get_count_details(count_id)
    
    # 建立 Excel 工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "盤點清單"
    
    # 定義樣式
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 寫入標題行
    headers = ['零件編號', '零件名稱', '單位', '系統數量', '實盤數量', '備註']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 寫入資料行
    for row_num, detail in enumerate(details, 2):
        # 零件編號
        cell = ws.cell(row=row_num, column=1)
        cell.value = detail['part_number']
        cell.border = border
        
        # 零件名稱
        cell = ws.cell(row=row_num, column=2)
        cell.value = detail['part_name']
        cell.border = border
        
        # 單位
        cell = ws.cell(row=row_num, column=3)
        cell.value = detail['unit']
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # 系統數量
        cell = ws.cell(row=row_num, column=4)
        cell.value = detail['system_quantity']
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        
        # 實盤數量（空白，待填入）
        cell = ws.cell(row=row_num, column=5)
        cell.value = detail.get('counted_quantity', '')
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        # 設定黃色背景提示需要填寫
        if not detail.get('counted_quantity'):
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # 備註
        cell = ws.cell(row=row_num, column=6)
        cell.value = detail.get('notes', '')
        cell.border = border
    
    # 自動調整欄寬（考慮中文字元寬度）
    def get_cell_width(text):
        """計算儲存格寬度，中文字元算2個單位"""
        if not text:
            return 0
        width = 0
        for char in str(text):
            # 中文字元、全形符號
            if '\u4e00' <= char <= '\u9fff' or '\u3000' <= char <= '\u303f':
                width += 2
            else:
                width += 1
        return width
    
    for col_num in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_num)
        
        # 計算該欄最大寬度
        max_length = 0
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_width = get_cell_width(cell.value)
                    if cell_width > max_length:
                        max_length = cell_width
            except:
                pass
        
        # 設定欄寬（加一些邊距，並限制最大值）
        adjusted_width = min(max_length + 2, 60)  # 最大60字元
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 凍結首行
    ws.freeze_panes = 'A2'
    
    # 儲存到記憶體
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # 產生檔案名稱
    from urllib.parse import quote
    count_number = count_info.get('count_number', count_id)
    filename = f"盤點清單_{count_number}.xlsx"
    # URL 編碼檔名以支援中文
    encoded_filename = quote(filename)
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f"attachment; filename*=UTF-8''{encoded_filename}"
        }
    )
