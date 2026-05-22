"""
Service layer for handling report generation business logic.
"""
from collections import defaultdict
from extensions import db
from models.work_order import WorkOrderDemand
from models.inventory import CurrentInventory, InventoryTransaction
from models.part import Part, Warehouse, WarehouseLocation
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from utils.datetime_utils import get_taipei_time


class ReportService:

    @staticmethod
    def _get_idle_bucket(idle_days):
        if idle_days is None:
            return 'no_consumption_history'
        if idle_days >= 180:
            return 'obsolete'
        if idle_days >= 90:
            return 'stagnant'
        if idle_days >= 30:
            return 'aging'
        return 'normal'

    @staticmethod
    def _idle_sort_key(item):
        bucket_priority = {
            'no_consumption_history': 0,
            'obsolete': 1,
            'stagnant': 2,
            'aging': 3,
            'normal': 4,
        }
        idle_days = item['idle_days'] if item['idle_days'] is not None else 999999
        return (
            bucket_priority.get(item['idle_bucket'], 9),
            -idle_days,
            item['part_number'],
            item['warehouse_code'],
            item['location_code'],
        )

    @staticmethod
    def get_idle_inventory_report_data():
        """Generate location-level idle inventory report data based on last consumption date."""
        try:
            consumption_types = InventoryTransaction.get_consumption_transaction_types()

            last_consumption_subquery = db.session.query(
                InventoryTransaction.part_id.label('part_id'),
                InventoryTransaction.warehouse_location_id.label('warehouse_location_id'),
                db.func.max(InventoryTransaction.transaction_date).label('last_consumption_date')
            ).filter(
                InventoryTransaction.warehouse_location_id.isnot(None),
                InventoryTransaction.transaction_type.in_(consumption_types)
            ).group_by(
                InventoryTransaction.part_id,
                InventoryTransaction.warehouse_location_id
            ).subquery()

            rows = db.session.query(
                Part.part_number,
                Part.name.label('part_name'),
                Part.type.label('part_type'),
                Part.unit,
                Warehouse.id.label('warehouse_id'),
                Warehouse.code.label('warehouse_code'),
                Warehouse.name.label('warehouse_name'),
                WarehouseLocation.id.label('warehouse_location_id'),
                WarehouseLocation.location_code,
                db.func.coalesce(CurrentInventory.quantity_on_hand, 0).label('quantity_on_hand'),
                db.func.coalesce(CurrentInventory.available_quantity, 0).label('available_quantity'),
                db.func.coalesce(CurrentInventory.reserved_quantity, 0).label('reserved_quantity'),
                last_consumption_subquery.c.last_consumption_date
            ).join(
                Part, Part.id == CurrentInventory.part_id
            ).join(
                WarehouseLocation, WarehouseLocation.id == CurrentInventory.warehouse_location_id
            ).join(
                Warehouse, Warehouse.id == CurrentInventory.warehouse_id
            ).outerjoin(
                last_consumption_subquery,
                db.and_(
                    last_consumption_subquery.c.part_id == CurrentInventory.part_id,
                    last_consumption_subquery.c.warehouse_location_id == CurrentInventory.warehouse_location_id,
                )
            ).filter(
                db.func.coalesce(CurrentInventory.quantity_on_hand, 0) > 0
            ).all()

            today = get_taipei_time().date()
            items = []
            warehouse_options = {}
            unique_part_numbers = set()
            total_quantity = 0
            over_30 = 0
            over_60 = 0
            over_90 = 0
            over_180 = 0
            no_history = 0

            for row in rows:
                last_consumption_date = row.last_consumption_date.isoformat() if row.last_consumption_date else None
                idle_days = None
                if row.last_consumption_date:
                    idle_days = max(0, (today - row.last_consumption_date.date()).days)
                idle_bucket = ReportService._get_idle_bucket(idle_days)

                if idle_days is None:
                    no_history += 1
                else:
                    if idle_days >= 30:
                        over_30 += 1
                    if idle_days >= 60:
                        over_60 += 1
                    if idle_days >= 90:
                        over_90 += 1
                    if idle_days >= 180:
                        over_180 += 1

                quantity_on_hand = int(row.quantity_on_hand or 0)
                available_quantity = int(row.available_quantity or 0)
                reserved_quantity = int(row.reserved_quantity or 0)

                items.append({
                    'part_number': row.part_number,
                    'part_name': row.part_name,
                    'part_type': row.part_type or '',
                    'unit': row.unit or '',
                    'warehouse_id': row.warehouse_id,
                    'warehouse_code': row.warehouse_code,
                    'warehouse_name': row.warehouse_name,
                    'warehouse_location_id': row.warehouse_location_id,
                    'location_code': row.location_code,
                    'quantity_on_hand': quantity_on_hand,
                    'available_quantity': available_quantity,
                    'reserved_quantity': reserved_quantity,
                    'last_consumption_date': last_consumption_date,
                    'idle_days': idle_days,
                    'idle_bucket': idle_bucket,
                })

                warehouse_options[row.warehouse_id] = {
                    'id': row.warehouse_id,
                    'code': row.warehouse_code,
                    'name': row.warehouse_name,
                }
                unique_part_numbers.add(row.part_number)
                total_quantity += quantity_on_hand

            items.sort(key=ReportService._idle_sort_key)

            return {
                'success': True,
                'summary': {
                    'total_locations': len(items),
                    'total_part_numbers': len(unique_part_numbers),
                    'total_quantity': total_quantity,
                    'idle_over_30_count': over_30,
                    'idle_over_60_count': over_60,
                    'idle_over_90_count': over_90,
                    'idle_over_180_count': over_180,
                    'no_consumption_history_count': no_history,
                },
                'items': items,
                'filters': {
                    'warehouses': sorted(warehouse_options.values(), key=lambda w: (w['code'], w['name']))
                }
            }
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def export_idle_inventory_excel():
        """Export idle inventory report data as an Excel file."""
        report_data = ReportService.get_idle_inventory_report_data()

        if not report_data['success']:
            raise Exception(report_data['error'])

        wb = Workbook()
        ws = wb.active
        ws.title = '閒置庫存分析'

        headers = [
            '零件編號', '零件名稱', '類型', '倉庫', '儲位', '在庫數量', '可用數量',
            '預留數量', '單位', '最後耗用日', '閒置天數', '閒置分級'
        ]
        ws.append(headers)

        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        bucket_labels = {
            'no_consumption_history': '上線後未領料',
            'obsolete': '180+ 天',
            'stagnant': '90+ 天',
            'aging': '30+ 天',
            'normal': '30 天內',
        }

        for item in report_data['items']:
            ws.append([
                item['part_number'],
                item['part_name'],
                item['part_type'],
                f"{item['warehouse_name']} ({item['warehouse_code']})",
                item['location_code'],
                item['quantity_on_hand'],
                item['available_quantity'],
                item['reserved_quantity'],
                item['unit'],
                item['last_consumption_date'] or '上線後未領料',
                item['idle_days'] if item['idle_days'] is not None else '上線後未領料',
                bucket_labels.get(item['idle_bucket'], item['idle_bucket'])
            ])

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        filename = f'閒置庫存分析_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return output.getvalue(), filename

    @staticmethod
    def get_parts_comparison_data():
        """
        Generates the parts comparison report data.
        This logic was moved from web_controller._get_parts_comparison_report_data
        """
        try:
            # 1. 取得所有工單需求
            all_demands = WorkOrderDemand.query.all()
            work_order_dict = defaultdict(lambda: {
                'description': '',
                'total_required': 0,
                'order_ids': set(),
                'names': set()
            })

            for demand in all_demands:
                part_number = demand.part_number
                work_order_dict[part_number]['description'] = demand.material_description
                work_order_dict[part_number]['total_required'] += demand.required_quantity
                work_order_dict[part_number]['order_ids'].add(demand.order_id)
                part_from_db = Part.query.filter_by(part_number=part_number).first()
                if part_from_db:
                    work_order_dict[part_number]['names'].add(part_from_db.name)
                else:
                    work_order_dict[part_number]['names'].add(demand.material_description)

            # 2. 取得零件倉零件的詳細資訊
            inventory_parts = db.session.query(Part).options(db.joinedload(Part.location_associations)).all()
            inventory_dict = {part.part_number: part for part in inventory_parts}

            # 3. 取得零件庫存資訊
            stock_details = db.session.query(
                Part.part_number,
                db.func.sum(CurrentInventory.quantity_on_hand).label('total_stock'),
                db.func.sum(CurrentInventory.available_quantity).label('available_stock')
            ).join(CurrentInventory, Part.id == CurrentInventory.part_id).group_by(Part.part_number).all()
            stock_dict = {row[0]: {'total_stock': float(row[1] or 0), 'available_stock': float(row[2] or 0)} for row in stock_details}

            # 4. 分析差異
            missing_in_inventory = []
            for part_number, details in work_order_dict.items():
                if part_number not in inventory_dict:
                    missing_in_inventory.append({
                        'part_number': part_number,
                        'description': details['description'],
                        'name': list(details['names'])[0] if details['names'] else details['description'],
                        'total_required': details['total_required'],
                        'order_ids': list(details['order_ids'])
                    })

            demand_with_no_location = []
            for part_number, details in work_order_dict.items():
                if part_number in inventory_dict:
                    part_obj = inventory_dict[part_number]
                    if not part_obj.location_associations:
                        demand_with_no_location.append({
                            'part_id': part_obj.id,
                            'part_number': part_number,
                            'name': part_obj.name,
                            'total_required': details['total_required'],
                            'order_ids': list(details['order_ids'])
                        })

            inventory_with_demand = []
            for part_number, part in inventory_dict.items():
                work_order_info = work_order_dict.get(part_number, {})
                stock_info = stock_dict.get(part_number, {'total_stock': 0, 'available_stock': 0})
                required_qty = work_order_info.get('total_required', 0)
                available_qty = stock_info.get('available_stock', 0)
                shortage = max(0, required_qty - available_qty)
                inventory_with_demand.append({
                    'part_number': part_number,
                    'name': part.name,
                    'description': part.description or '',
                    'unit': part.unit or '',
                    'category': part.type or '',
                    'warehouse_location_id': part.location_associations[0].warehouse_location_id if part.location_associations else None,
                    'required_quantity': required_qty,
                    'total_stock': stock_info.get('total_stock', 0),
                    'available_quantity': available_qty,
                    'shortage': shortage,
                    'order_ids': list(work_order_info.get('order_ids', [])),
                    'has_demand': required_qty > 0,
                    'stock_status': '充足' if shortage == 0 and required_qty > 0 else ('缺料' if shortage > 0 else '無需求')
                })

            # 5. 計算統計資訊
            summary = {
                'work_order_parts_count': len(work_order_dict),
                'inventory_parts_count': len(inventory_dict),
                'missing_in_inventory_count': len(missing_in_inventory),
                'demand_with_no_location_count': len(demand_with_no_location),
                'shortage_parts_count': len([item for item in inventory_with_demand if item['shortage'] > 0]),
                'sufficient_parts_count': len([item for item in inventory_with_demand if item['stock_status'] == '充足'])
            }

            return {
                'success': True,
                'summary': summary,
                'missing_in_inventory': sorted(missing_in_inventory, key=lambda x: x['total_required'], reverse=True),
                'inventory_with_demand': sorted(inventory_with_demand, key=lambda x: x['shortage'], reverse=True),
                'demand_with_no_location': sorted(demand_with_no_location, key=lambda x: x['total_required'], reverse=True)
            }

        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': str(e)
            }

    @staticmethod
    def export_parts_comparison_excel():
        """
        Generates and exports the parts comparison report as an Excel file.
        This logic was moved from web_controller.export_parts_comparison.
        """
        report_data = ReportService.get_parts_comparison_data()

        if not report_data['success']:
            # Propagate the error to be handled by the controller
            raise Exception(report_data['error'])

        wb = Workbook()

        # Define styles
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill_missing = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        header_fill_shortage = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        header_fill_sufficient = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
        header_fill_no_location = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        def add_sheet_data(ws, title, headers, data, header_fill):
            ws.title = title
            ws.append(headers)
            for col_idx, cell in enumerate(ws[1]):
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                ws.column_dimensions[get_column_letter(col_idx + 1)].width = 15

            for row_data in data:
                ws.append(list(row_data.values()))
                for cell in ws[ws.max_row]:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='left', vertical='center')
            
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column].width = adjusted_width

        # Sheet 1: Missing in Inventory
        missing_data_raw = report_data['missing_in_inventory']
        if missing_data_raw:
            missing_data_transformed = []
            for item in missing_data_raw:
                missing_data_transformed.append({
                    'part_number': item['part_number'],
                    'name': item['name'],
                    'total_required': item['total_required'],
                    'order_ids': ', '.join(item['order_ids'])
                })
            ws1 = wb.active
            add_sheet_data(ws1, '未建立零件', ['零件編號', '零件名稱', '需求數量', '工單號碼'], missing_data_transformed, header_fill_missing)
        else:
            ws1 = wb.active
            ws1.title = '未建立零件'
            ws1.append(['無數據'])

        # Sheet 2: Shortage Parts
        shortage_data = [item for item in report_data['inventory_with_demand'] if item['shortage'] > 0]
        if shortage_data:
            ws2 = wb.create_sheet('庫存不足')
            add_sheet_data(ws2, '庫存不足', ['零件編號', '零件名稱', '需求數量', '庫存數量', '缺貨數量', '單位', '工單號碼'],
                           [{ 'part_number': item['part_number'], 'name': item['name'], 'required_quantity': item['required_quantity'],
                             'total_stock': item['total_stock'], 'shortage': item['shortage'], 'unit': item['unit'],
                             'order_ids': ', '.join(item['order_ids'])} for item in shortage_data],
                           header_fill_shortage)
        else:
            ws2 = wb.create_sheet('庫存不足')
            ws2.append(['無數據'])

        # Sheet 3: Sufficient Parts
        sufficient_data = [item for item in report_data['inventory_with_demand'] if item['stock_status'] == '充足']
        if sufficient_data:
            ws3 = wb.create_sheet('庫存充足')
            add_sheet_data(ws3, '庫存充足', ['零件編號', '零件名稱', '需求數量', '庫存數量', '可用數量', '單位', '工單號碼'],
                           [{ 'part_number': item['part_number'], 'name': item['name'], 'required_quantity': item['required_quantity'],
                             'total_stock': item['total_stock'], 'available_quantity': item['available_quantity'], 'unit': item['unit'],
                             'order_ids': ', '.join(item['order_ids'])} for item in sufficient_data],
                           header_fill_sufficient)
        else:
            ws3 = wb.create_sheet('庫存充足')
            ws3.append(['無數據'])

        # Sheet 4: Demand with No Location
        no_location_data = report_data['demand_with_no_location']
        if no_location_data:
            ws4 = wb.create_sheet('待建立儲位')
            add_sheet_data(ws4, '待建立儲位', ['零件編號', '零件名稱', '總需求數量', '工單號碼'],
                           [{ 'part_number': item['part_number'], 'name': item['name'], 'total_required': item['total_required'],
                             'order_ids': ', '.join(item['order_ids'])} for item in no_location_data],
                           header_fill_no_location)
        else:
            ws4 = wb.create_sheet('待建立儲位')
            ws4.append(['無數據'])

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        filename = f'零件差異分析報告_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return output.getvalue(), filename
