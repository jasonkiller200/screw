from extensions import db
from models.weekly_order import OrderRegistration, WeeklyOrderCycle, OrderReviewLog
from models.part import Part
import pandas as pd
from io import BytesIO
from datetime import datetime

class WeeklyOrderService:
    """
    Service layer for handling weekly order related business logic.
    """

    @staticmethod
    def export_weekly_order_excel(cycle_id):
        from models.weekly_order import WeeklyOrderCycle, OrderRegistration, OrderReviewLog
        from models.part import WarehouseLocation
        from sqlalchemy.orm import joinedload
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        try:
            cycle = WeeklyOrderCycle.query.get(cycle_id)
            if not cycle:
                return {'success': False, 'message': '找不到指定的週期'}

            registrations = OrderRegistration.query.options(
                joinedload(OrderRegistration.warehouse_location).joinedload(WarehouseLocation.warehouse)
            ).filter(
                OrderRegistration.cycle_id == cycle_id, 
                OrderRegistration.status == 'approved'
            ).order_by(OrderRegistration.item_sequence).all()
            
            if not registrations:
                return {'success': False, 'message': '沒有已核准的項目可生成申請單'}
            
            # 獲取審查記錄
            review_logs = OrderReviewLog.query.filter_by(
                cycle_id=cycle_id,
                action='approved'
            ).order_by(OrderReviewLog.created_at.desc()).all()
            
            reviewers = set()
            for log in review_logs:
                if log.reviewer_name and log.reviewer_name not in ['主管', '系統']:
                    reviewers.add(log.reviewer_name)
            reviewers_str = '、'.join(sorted(reviewers)) if reviewers else '系統'
            
            # 獲取申請單位（從登記項目中取得）
            departments = set()
            for reg in registrations:
                if reg.department and reg.department != '未設定':
                    departments.add(reg.department)
            application_unit = '、'.join(sorted(departments)) if departments else '生產部'
            
            # Helper function to get current time in UTC+8
            def get_taipei_time():
                from datetime import timezone, timedelta
                tz_taipei = timezone(timedelta(hours=8))
                return datetime.now(tz_taipei)
            
            # 建立 Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = '採購申請單'
            
            # 定義樣式
            title_font = Font(name='微軟正黑體', size=16, bold=True)
            header_font = Font(name='微軟正黑體', size=11, bold=True, color='FFFFFF')
            normal_font = Font(name='微軟正黑體', size=10)
            
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            
            center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            # Row 1: 標題
            ws.merge_cells('A1:L1')
            ws['A1'] = 'Hartford螺絲/接頭/五金/耗材用品申請'
            ws['A1'].font = title_font
            ws['A1'].alignment = center_alignment
            ws.row_dimensions[1].height = 30
            
            # Row 2: 自我管理提醒
            ws.merge_cells('A2:L2')
            ws['A2'] = '*自我管理:請購之前先檢討是否真有必要?如非買不可.在予申請'
            ws['A2'].font = Font(name='微軟正黑體', size=11, bold=True, color='FF0000')  # 紅色字體
            ws['A2'].alignment = center_alignment
            ws.row_dimensions[2].height = 25
            
            # Row 3: 申請日期與申請單位
            ws.merge_cells('A3:F3')
            ws['A3'] = f'申請單位：{application_unit}' 
            ws['A3'].font = Font(name='微軟正黑體', size=11)
            ws['A3'].alignment = left_alignment
            
            ws.merge_cells('G3:L3')
            ws['G3'] = f'申請日期：{get_taipei_time().strftime("%Y-%m-%d")}'
            ws['G3'].font = Font(name='微軟正黑體', size=11)
            ws['G3'].alignment = left_alignment
            
            # Row 4: 請購類型
            ws.merge_cells('A4:L4')
            ws['A4'] = '□一般存貨請購(非使用於工單)■事務性請購(使用於工單)'
            ws['A4'].font = Font(name='微軟正黑體', size=11, bold=True)
            ws['A4'].alignment = left_alignment
            ws.row_dimensions[4].height = 25
            
            # 空白行
            ws.row_dimensions[5].height = 5
            
            # 表頭（第6行）
            headers = ['項次', '品號', '品名', '儲位', '種類', '數量', '單位', 
                      '申請人', '申請單位', '緊急程度', '需用日期', '台份用/備註']
            
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=6, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border
            
            ws.row_dimensions[6].height = 25
            
            # 資料列（從第7行開始）
            for idx, reg in enumerate(registrations, start=1):
                row_num = 6 + idx
                
                location_str = ''
                if reg.warehouse_location and reg.warehouse_location.warehouse:
                    location_str = f"{reg.warehouse_location.warehouse.name} - {reg.warehouse_location.location_code}"
                elif reg.warehouse_location:
                    location_str = reg.warehouse_location.location_code
                
                priority_str = '緊急' if reg.priority == 'urgent' else '一般'
                
                row_data = [
                    idx,
                    reg.part_number,
                    reg.part_name,
                    location_str,
                    reg.category or '',
                    reg.quantity,
                    reg.unit,
                    reg.applicant_name,
                    reg.department or '',
                    priority_str,
                    reg.required_date.strftime('%Y-%m-%d') if reg.required_date else '',
                    reg.purpose_notes or ''
                ]
                
                for col_idx, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.value = value
                    cell.font = normal_font
                    cell.border = thin_border
                    
                    # 對齊方式
                    if col_idx in [1, 6]:  # 項次、數量
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
                
                ws.row_dimensions[row_num].height = 20
            
            # 添加注意事項（在所有資料列之後）
            last_data_row = 6 + len(registrations)
            
            # 空白行
            ws.row_dimensions[last_data_row + 1].height = 15
            
            # 注意事項標題
            notes_start_row = last_data_row + 2
            ws.merge_cells(f'A{notes_start_row}:L{notes_start_row}')
            ws[f'A{notes_start_row}'] = '注意事項：'
            ws[f'A{notes_start_row}'].font = Font(name='微軟正黑體', size=11, bold=True)
            ws[f'A{notes_start_row}'].alignment = left_alignment
            ws.row_dimensions[notes_start_row].height = 20
            
            # 注意事項內容
            note1_row = notes_start_row + 1
            ws.merge_cells(f'A{note1_row}:L{note1_row}')
            ws[f'A{note1_row}'] = '1.請每週提出請購。'
            ws[f'A{note1_row}'].font = normal_font
            ws[f'A{note1_row}'].alignment = left_alignment
            ws.row_dimensions[note1_row].height = 20
            
            note2_row = notes_start_row + 2
            ws.merge_cells(f'A{note2_row}:L{note2_row}')
            ws[f'A{note2_row}'] = '2.一次性,費用較高者,請供應商先提供估價單。'
            ws[f'A{note2_row}'].font = normal_font
            ws[f'A{note2_row}'].alignment = left_alignment
            ws.row_dimensions[note2_row].height = 20
            
            note3_row = notes_start_row + 3
            ws.merge_cells(f'A{note3_row}:L{note3_row}')
            ws[f'A{note3_row}'] = '3.申請經核准始可購買,不得先行購買。'
            ws[f'A{note3_row}'].font = normal_font
            ws[f'A{note3_row}'].alignment = left_alignment
            ws.row_dimensions[note3_row].height = 20
            
            # 單位主管印章區域（在注意事項之後）
            reviewer_row = note3_row + 2
            
            # 用於追蹤需要清理的臨時文件
            temp_files_to_clean = []
            
            # 左側：單位主管標籤（不顯示審查者姓名）
            ws.merge_cells(f'A{reviewer_row}:B{reviewer_row}')
            ws[f'A{reviewer_row}'] = '單位主管'
            ws[f'A{reviewer_row}'].font = Font(name='微軟正黑體', size=12, bold=True)
            ws[f'A{reviewer_row}'].alignment = center_alignment
            
            # 嘗試插入印章圖片
            try:
                from openpyxl.drawing.image import Image as XLImage
                import os
                
                # 直接生成動態印章（使用審查者姓名）
                stamp_image_path = WeeklyOrderService._generate_stamp_image(reviewers_str)
                if stamp_image_path:
                    temp_files_to_clean.append(stamp_image_path)  # 記錄臨時文件
                    try:
                        stamp_img = XLImage(stamp_image_path)
                        stamp_img.width = 100
                        stamp_img.height = 60
                        # 將印章放在 C 欄位（單位主管標籤旁邊）
                        stamp_img.anchor = f'C{reviewer_row}'
                        ws.add_image(stamp_img)
                        ws.row_dimensions[reviewer_row].height = 50
                        
                        # 在印章旁邊顯示日期標籤和日期
                        ws.merge_cells(f'E{reviewer_row}:F{reviewer_row}')
                        ws[f'E{reviewer_row}'] = '核准日期'
                        ws[f'E{reviewer_row}'].font = Font(name='微軟正黑體', size=12, bold=True)
                        ws[f'E{reviewer_row}'].alignment = center_alignment
                        
                        ws.merge_cells(f'G{reviewer_row}:H{reviewer_row}')
                        ws[f'G{reviewer_row}'] = get_taipei_time().strftime("%Y/%m/%d")
                        ws[f'G{reviewer_row}'].font = Font(name='微軟正黑體', size=10, bold=True)
                        ws[f'G{reviewer_row}'].alignment = center_alignment
                    except Exception as img_error:
                        print(f"❌ Debug: 印章插入失敗 = {img_error}")
                        ws.row_dimensions[reviewer_row].height = 25
                else:
                    # 如果印章生成失敗，保持空白
                    ws.row_dimensions[reviewer_row].height = 25
                        
            except ImportError:
                # 如果無法匯入圖片模組，跳過印章區域
                ws.row_dimensions[reviewer_row].height = 25
            
            # 設定欄寬
            column_widths = {
                'A': 8,   # 項次
                'B': 15,  # 品號
                'C': 25,  # 品名
                'D': 20,  # 儲位
                'E': 12,  # 種類
                'F': 10,  # 數量
                'G': 8,   # 單位
                'H': 12,  # 申請人
                'I': 15,  # 申請單位
                'J': 12,  # 緊急程度
                'K': 12,  # 需用日期
                'L': 30   # 台份用/備註
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # 儲存到 BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # 清理臨時文件（在工作簿保存之後）
            for temp_file in temp_files_to_clean:
                try:
                    os.remove(temp_file)
                except:
                    pass

            filename = f"Hartford螺絲五金耗材用品申請_{get_taipei_time().strftime('%Y%m%d')}.xlsx"
            
            # 標記 Excel 已生成（不修改週期狀態，保持 reviewing）
            cycle.excel_generated = True
            cycle.reviewed_at = get_taipei_time()
            
            review_log = OrderReviewLog(
                cycle_id=cycle.id,
                reviewer_name='系統',
                action='export_excel',
                notes=f'匯出Excel申請單，包含{len(registrations)}個項目'
            )
            db.session.add(review_log)
            db.session.commit()
            
            return {'success': True, 'file_content': output.getvalue(), 'filename': filename}
            
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f"匯出失敗: {str(e)}"}

    @staticmethod
    def _generate_stamp_image(reviewer_name):
        """動態生成印章圖片，返回臨時檔案路徑"""
        try:
            from PIL import Image, ImageDraw, ImageFont
            import os
            import tempfile
            
            # 創建長方形印章 (100x60 像素)
            size = (100, 60)
            
            # 創建透明背景的圖片
            img = Image.new('RGBA', size, (255, 255, 255, 0))
            draw = ImageDraw.Draw(img)
            
            # 印章顏色（紅色）
            stamp_color = (220, 0, 0, 255)  # 紅色
            
            # 畫外框邊框（長方形）
            draw.rectangle([2, 2, 98, 58], outline=stamp_color, width=3)
            
            # 畫內框邊框（長方形）
            draw.rectangle([6, 6, 94, 54], outline=stamp_color, width=2)
            
            # 嘗試載入字體
            try:
                # Windows 系統字體路徑
                font_paths = [
                    'C:/Windows/Fonts/kaiu.ttf',  # 標楷體
                    'C:/Windows/Fonts/simhei.ttf',  # 黑體
                    'C:/Windows/Fonts/msyh.ttc',   # 微軟雅黑
                    '/System/Library/Fonts/STKaiti.ttc',  # Mac 標楷體
                    '/usr/share/fonts/truetype/arphic/ukai.ttc'  # Linux
                ]
                
                font = None
                for font_path in font_paths:
                    if os.path.exists(font_path):
                        font = ImageFont.truetype(font_path, 24)  # 使用 24pt 字體
                        break
                
                if not font:
                    font = ImageFont.load_default()
                    
            except Exception:
                font = ImageFont.load_default()
            
            # 計算文字位置（居中）
            text = reviewer_name
            if len(text) > 6:
                text = text[:6]  # 長方形印章可容納更多字元
            
            # 取得文字邊界 (使用 textbbox 替代已棄用的 textsize)
            bbox = draw.textbbox((0, 0), text, font=font)
            text_width = bbox[2] - bbox[0]
            text_height = bbox[3] - bbox[1]
            
            # 計算居中位置
            x = (size[0] - text_width) // 2
            y = (size[1] - text_height) // 2
            
            # 畫文字
            draw.text((x, y), text, fill=stamp_color, font=font)
            
            # 儲存到臨時檔案
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
            img.save(temp_file.name, 'PNG')
            temp_file.close()
            
            return temp_file.name
            
        except ImportError:
            # PIL 未安裝
            return None
        except Exception as e:
            # 其他錯誤
            print(f"生成印章圖片失敗: {e}")
            return None

    @staticmethod
    def export_weekly_order_pdf(cycle_id):
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as ReportLabImage
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import mm
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        import io
        import os
        import tempfile
        from models.weekly_order import WeeklyOrderCycle, OrderRegistration, OrderReviewLog
        from models.part import WarehouseLocation
        from sqlalchemy.orm import joinedload

        try:
            cycle = WeeklyOrderCycle.query.get(cycle_id)
            if not cycle:
                return {'success': False, 'message': '找不到指定的週期'}

            registrations = OrderRegistration.query.options(
                joinedload(OrderRegistration.warehouse_location).joinedload(WarehouseLocation.warehouse)
            ).filter(
                OrderRegistration.cycle_id == cycle_id,
                OrderRegistration.status == 'approved'
            ).order_by(OrderRegistration.item_sequence).all()

            if not registrations:
                return {'success': False, 'message': '沒有已核准的項目可生成 PDF'}

            # 獲取審查記錄
            review_logs = OrderReviewLog.query.filter_by(
                cycle_id=cycle_id,
                action='approved'
            ).order_by(OrderReviewLog.created_at.desc()).all()
            
            reviewers = set()
            for log in review_logs:
                if log.reviewer_name and log.reviewer_name not in ['主管', '系統']:
                    reviewers.add(log.reviewer_name)
            reviewers_str = '、'.join(sorted(reviewers)) if reviewers else '系統'
            
            # 獲取申請單位
            departments = set()
            for reg in registrations:
                if reg.department and reg.department != '未設定':
                    departments.add(reg.department)
            application_unit = '、'.join(sorted(departments)) if departments else '生產部'

            # Helper function to get current time in UTC+8
            def get_taipei_time():
                from datetime import timezone, timedelta
                tz_taipei = timezone(timedelta(hours=8))
                return datetime.now(tz_taipei)

            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=landscape(A4), 
                rightMargin=15*mm, 
                leftMargin=15*mm, 
                topMargin=15*mm, 
                bottomMargin=15*mm
            )
            
            # 註冊中文字體
            font_path = "C:/Windows/Fonts/kaiu.ttf"
            if not os.path.exists(font_path):
                font_path = "C:/Windows/Fonts/simhei.ttf"
            
            if os.path.exists(font_path):
                pdfmetrics.registerFont(TTFont('ChineseFont', font_path))
                font_name = 'ChineseFont'
            else:
                font_name = 'Helvetica'

            elements = []
            
            # 定義樣式
            title_style = ParagraphStyle(
                'CustomTitle',
                fontName=font_name,
                fontSize=16,
                alignment=TA_CENTER,
                spaceAfter=6,
                leading=20
            )
            
            warning_style = ParagraphStyle(
                'Warning',
                fontName=font_name,
                fontSize=11,
                alignment=TA_CENTER,
                textColor=colors.red,
                spaceAfter=6,
                leading=14
            )
            
            info_style = ParagraphStyle(
                'Info',
                fontName=font_name,
                fontSize=10,
                alignment=TA_LEFT,
                spaceAfter=3,
                leading=12
            )
            
            note_style = ParagraphStyle(
                'Note',
                fontName=font_name,
                fontSize=9,
                alignment=TA_LEFT,
                spaceAfter=3,
                leading=11
            )
            
            # 標題
            elements.append(Paragraph('Hartford螺絲/接頭/五金/耗材用品申請', title_style))
            
            # 自我管理提醒
            elements.append(Paragraph('*自我管理:請購之前先檢討是否真有必要?如非買不可.在予申請', warning_style))
            
            # 申請單位和日期
            info_table_data = [
                [Paragraph(f'申請單位：{application_unit}', info_style), 
                 Paragraph(f'申請日期：{get_taipei_time().strftime("%Y-%m-%d")}', info_style)]
            ]
            info_table = Table(info_table_data, colWidths=[140*mm, 100*mm])
            info_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            elements.append(info_table)
            
            # 請購類型
            elements.append(Paragraph('□一般存貨請購(非使用於工單)■事務性請購(使用於工單)', info_style))
            elements.append(Spacer(1, 3*mm))
            
            # 主表格資料
            headers = ['項次', '品號', '品名', '儲位', '種類', '數量', '單位', 
                      '申請人', '申請單位', '緊急程度', '需用日期', '台份用/備註']
            data = [headers]
            
            for idx, reg in enumerate(registrations, start=1):
                location_str = ''
                if reg.warehouse_location and reg.warehouse_location.warehouse:
                    location_str = f"{reg.warehouse_location.warehouse.name}-{reg.warehouse_location.location_code}"
                elif reg.warehouse_location:
                    location_str = reg.warehouse_location.location_code
                
                priority_str = '緊急' if reg.priority == 'urgent' else '一般'
                
                row = [
                    str(idx),
                    reg.part_number or '',
                    Paragraph(reg.part_name or '', note_style),
                    location_str,
                    reg.category or '',
                    str(reg.quantity),
                    reg.unit or '',
                    reg.applicant_name or '',
                    reg.department or '',
                    priority_str,
                    reg.required_date.strftime('%Y-%m-%d') if reg.required_date else '',
                    Paragraph(reg.purpose_notes or '', note_style)
                ]
                data.append(row)
            
            # 主表格
            main_table = Table(data, colWidths=[
                10*mm,  # 項次
                20*mm,  # 品號
                35*mm,  # 品名
                25*mm,  # 儲位
                15*mm,  # 種類
                12*mm,  # 數量
                10*mm,  # 單位
                15*mm,  # 申請人
                18*mm,  # 申請單位
                15*mm,  # 緊急程度
                18*mm,  # 需用日期
                40*mm   # 台份用/備註
            ])
            
            main_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F0F0')]),
            ]))
            
            elements.append(main_table)
            elements.append(Spacer(1, 5*mm))
            
            # 注意事項
            elements.append(Paragraph('<b>注意事項：</b>', info_style))
            elements.append(Paragraph('1.請每週提出請購。', note_style))
            elements.append(Paragraph('2.一次性,費用較高者,請供應商先提供估價單。', note_style))
            elements.append(Paragraph('3.申請經核准始可購買,不得先行購買。', note_style))
            elements.append(Spacer(1, 5*mm))
            
            # 單位主管印章區域
            stamp_data = []
            temp_files = []
            
            # 為印章區域標題創建置中樣式
            stamp_header_style = ParagraphStyle(
                'StampHeader',
                fontName=font_name,
                fontSize=10,
                alignment=TA_CENTER,
                spaceAfter=3,
                leading=12
            )
            
            # 嘗試生成印章
            stamp_image = None
            stamp_path = WeeklyOrderService._generate_stamp_image(reviewers_str)
            if stamp_path:
                temp_files.append(stamp_path)
                stamp_image = ReportLabImage(stamp_path, width=25*mm, height=15*mm)
            
            # 第一行：標題框
            header_row = [
                Paragraph('<b>單位主管</b>', stamp_header_style),
                Paragraph('<b>核准日期</b>', stamp_header_style)
            ]
            stamp_data.append(header_row)
            
            # 第二行：印章和日期
            if stamp_image:
                content_row = [
                    stamp_image,
                    Paragraph(f'{get_taipei_time().strftime("%Y/%m/%d")}', stamp_header_style)
                ]
            else:
                content_row = [
                    '',
                    Paragraph(f'{get_taipei_time().strftime("%Y/%m/%d")}', stamp_header_style)
                ]
            stamp_data.append(content_row)
            
            stamp_table = Table(stamp_data, colWidths=[40*mm, 40*mm])
            stamp_table.setStyle(TableStyle([
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                # 標題行樣式 - 白色背景，黑色文字
                ('BACKGROUND', (0, 0), (-1, 0), colors.white),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                # 內容行樣式
                ('ALIGN', (0, 1), (-1, 1), 'CENTER'),
                ('VALIGN', (0, 1), (-1, 1), 'MIDDLE'),
                # 邊框
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ]))
            
            elements.append(stamp_table)
            
            # 建立 PDF
            doc.build(elements)
            
            # 清理暫存檔
            for f in temp_files:
                try:
                    os.remove(f)
                except:
                    pass

            filename = f"Hartford螺絲五金耗材用品申請_{get_taipei_time().strftime('%Y%m%d')}.pdf"
            
            return {
                'success': True,
                'file_content': buffer.getvalue(),
                'filename': filename
            }
        except Exception as e:
            import traceback
            traceback.print_exc()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def review_order_registration(registration_id, action, notes, reviewer_id, reviewer_name):
        from models.weekly_order import OrderReviewLog
        from services.notification_service import NotificationService
        try:
            registration = OrderRegistration.query.get(registration_id)
            if not registration:
                return {'success': False, 'message': '找不到指定的登記項目'}
            
            old_status = registration.status
            
            if action == 'approved':
                registration.status = 'approved'
            elif action == 'rejected':
                registration.status = 'rejected'
            else:
                return {'success': False, 'message': '無效的操作'}
            
            registration.admin_notes = notes
            
            review_log = OrderReviewLog(
                cycle_id=registration.cycle_id,
                registration_id=registration.id,
                reviewer_id=reviewer_id,
                reviewer_name=reviewer_name,
                action=action,
                old_status=old_status,
                new_status=registration.status,
                notes=notes
            )
            db.session.add(review_log)
            
            # 如果是拒絕操作，且申請者有效，創建通知
            if action == 'rejected' and registration.applicant_id:
                title = f"週期訂單申請被拒絕"
                content = f"您的申請項目「{registration.part_name}」（品號：{registration.part_number}）已被拒絕。"
                if notes:
                    content += f"\n\n拒絕原因：{notes}"
                
                notification_success, notification_result = NotificationService.create_notification(
                    user_id=registration.applicant_id,
                    notification_type='order_rejected',
                    title=title,
                    content=content,
                    order_registration_id=registration.id
                )
                
                if not notification_success:
                    print(f"警告：創建拒絕通知失敗 - {notification_result}")
            
            db.session.commit()
            
            return {
                'success': True,
                'message': f"項目已{'通過' if action == 'approved' else '拒絕'}",
                'new_status': registration.status
            }
            
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def batch_review_registrations(registration_ids, action, reviewer_id, reviewer_name):
        from models.weekly_order import OrderRegistration, OrderReviewLog
        from services.notification_service import NotificationService
        try:
            updated_count = 0
            notification_count = 0
            
            for reg_id in registration_ids:
                registration = OrderRegistration.query.get(reg_id)
                if registration and registration.status == 'registered':
                    old_status = registration.status
                    registration.status = action

                    review_log = OrderReviewLog(
                        cycle_id=registration.cycle_id,
                        registration_id=registration.id,
                        reviewer_id=reviewer_id,
                        reviewer_name=reviewer_name,
                        action=action,
                        old_status=old_status,
                        new_status=registration.status,
                        notes=f'批量{action}'
                    )
                    db.session.add(review_log)
                    updated_count += 1

                    # 如果是拒絕操作，且申請者有效，創建通知
                    if action == 'rejected' and registration.applicant_id:
                        title = f"週期訂單申請被拒絕"
                        content = f"您的申請項目「{registration.part_name}」（品號：{registration.part_number}）在批量審查中被拒絕。"
                        
                        notification_success, notification_result = NotificationService.create_notification(
                            user_id=registration.applicant_id,
                            notification_type='order_rejected',
                            title=title,
                            content=content,
                            order_registration_id=registration.id
                        )
                        
                        if notification_success:
                            notification_count += 1
                        else:
                            print(f"警告：創建拒絕通知失敗 - {notification_result}")

            db.session.commit()

            message = f'已批量處理 {updated_count} 個項目'
            if action == 'rejected' and notification_count > 0:
                message += f'，已發送 {notification_count} 個拒絕通知'

            return {
                'success': True,
                'message': message
            }

        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': str(e)}

    @staticmethod
    def register_new_order(cycle_id, form_data, user_id, user_name, user_department):
        from models.weekly_order import WeeklyOrderCycle, OrderRegistration
        try:
            current_cycle = WeeklyOrderCycle.query.get(cycle_id)
            if not current_cycle:
                return {'success': False, 'message': '目前沒有活躍的申請週期'}

            if not current_cycle.is_active:
                return {'success': False, 'message': '申請週期已截止，無法新增登記'}

            part_number = form_data.get('part_number', '').strip()
            part = Part.query.filter_by(part_number=part_number).first()
            if part and not part.location_associations:
                if not form_data.get('purpose_notes', '').strip():
                    return {'success': False, 'message': '此零件無指定儲位，請務必填寫「台份用/備註」'}

            max_sequence = db.session.query(db.func.max(OrderRegistration.item_sequence)).filter_by(cycle_id=current_cycle.id).scalar()
            next_sequence = (max_sequence or 0) + 1
            
            warehouse_location_id = form_data.get('warehouse_location_id', type=int)

            registration = OrderRegistration(
                cycle_id=current_cycle.id,
                item_sequence=next_sequence,
                part_number=form_data.get('part_number', '').strip(),
                part_name=form_data.get('part_name', '').strip(),
                warehouse_location_id=warehouse_location_id,
                quantity=int(form_data.get('quantity', 0)),
                unit=form_data.get('unit', '').strip(),
                category=form_data.get('part_type', '').strip(),
                required_date=datetime.strptime(form_data.get('required_date'), '%Y-%m-%d') if form_data.get('required_date') else None,
                priority=form_data.get('priority', 'normal').strip(),
                purpose_notes=form_data.get('purpose_notes', '').strip(),
                applicant_id=user_id,
                applicant_name=user_name,
                department=user_department or '未設定'
            )
            
            db.session.add(registration)
            db.session.commit()
            
            return {'success': True, 'message': f'項目 #{next_sequence} 登記成功'}
            
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'登記失敗：{str(e)}'}

    @staticmethod
    def batch_register_orders(cycle_id, parts_data, source):
        from models.weekly_order import WeeklyOrderCycle, OrderRegistration
        try:
            current_cycle = WeeklyOrderCycle.query.get(cycle_id)
            if not current_cycle:
                return {'success': False, 'message': '目前沒有活躍的申請週期'}
            
            if not current_cycle.is_active:
                return {'success': False, 'message': '申請週期已截止，無法新增登記'}
            
            if not parts_data:
                return {'success': False, 'message': '沒有提供要登記的零件資料'}
            
            added_count = 0
            
            for part_data in parts_data:
                max_sequence = db.session.query(db.func.max(OrderRegistration.item_sequence)).filter_by(cycle_id=current_cycle.id).scalar()
                next_sequence = (max_sequence or 0) + 1
                
                warehouse_location_id = part_data.get('warehouse_location_id', type=int)

                registration = OrderRegistration()
                registration.cycle_id = current_cycle.id
                registration.item_sequence = next_sequence
                registration.part_number = part_data.get('part_number', '').strip()
                registration.part_name = part_data.get('part_name', '').strip()
                registration.warehouse_location_id = warehouse_location_id
                registration.quantity = int(part_data.get('quantity', 1))
                registration.unit = part_data.get('unit', '個').strip()
                registration.category = part_data.get('category', '').strip()
                registration.priority = part_data.get('priority', 'normal').strip()
                registration.purpose_notes = f'自動匯入自{source}'
                registration.applicant_name = '系統自動'
                registration.department = '自動申請'
                
                db.session.add(registration)
                added_count += 1
            
            db.session.commit()
            
            return {
                'success': True,
                'message': f'成功登記 {added_count} 個項目',
                'added_count': added_count
            }
            
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'批量登記失敗：{str(e)}'}

    @staticmethod
    def batch_inbound_weekly_orders(items):
        from models.weekly_order import OrderRegistration
        from services.inventory_service import InventoryService

        if not items or not isinstance(items, list):
            return {'success': False, 'error': '請求格式錯誤，需要一個項目列表'}

        success_count = 0
        error_count = 0
        errors = []

        for item in items:
            registration_id = item.get('registration_id')
            quantity = item.get('quantity')

            if not all([registration_id, quantity]):
                error_count += 1
                errors.append({'item': item, 'error': '缺少 registration_id 或 quantity'})
                continue
            
            try:
                inbound_quantity = int(quantity)
                if inbound_quantity <= 0:
                    error_count += 1
                    errors.append({'item': item, 'error': '入庫數量必須為正整數'})
                    continue
            except (ValueError, TypeError):
                error_count += 1
                errors.append({'item': item, 'error': '數量格式錯誤'})
                continue

            result = InventoryService.receive_stock(
                registration_id=registration_id, 
                inbound_quantity=inbound_quantity, 
                notes='批量入庫'
            )

            if result.get('success'):
                success_count += 1
            else:
                error_count += 1
                reg = OrderRegistration.query.get(registration_id)
                error_item_info = f"品號 {reg.part_number}" if reg else f"ID {registration_id}"
                errors.append({'item': error_item_info, 'error': result.get('error', '未知錯誤')})

        response = {
            'success': error_count == 0,
            'message': f'批量入庫完成。成功: {success_count} 項, 失敗: {error_count} 項。',
            'success_count': success_count,
            'error_count': error_count,
            'errors': errors
        }
        return response
