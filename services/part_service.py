import pandas as pd
from io import BytesIO
from models.part import Part, Warehouse, WarehouseLocation, PartWarehouseLocation
from extensions import db
from sqlalchemy.orm import joinedload
from models.weekly_order import OrderRegistration
from models.inventory import CurrentInventory
from models.part import WarehouseLocation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import or_, func

class PartService:

    @staticmethod
    def create_part_from_form(form_data):
        """Create a new part from submitted form data.

        Returns a dict with keys: success (bool), error (str) or data (form values)
        """
        try:
            part_number = form_data.get('part_number', '').strip()
            name = form_data.get('name', '').strip()
            type_ = form_data.get('type', '').strip()
            description = form_data.get('description', '').strip()
            unit = form_data.get('unit', '').strip() or '個'
            quantity_per_box = int(form_data.get('quantity_per_box') or 1)
            lead_time = int(form_data.get('lead_time') or 5)
            standard_cost = float(form_data.get('standard_cost') or 0)
            # desired_days_of_stock and moq are now per location, not part of top-level Part creation

            # Basic validation
            if not part_number:
                return {'success': False, 'error': '零件編號為必填項目', 'data': form_data}
            if not name:
                return {'success': False, 'error': '零件名稱為必填項目', 'data': form_data}

            # Parse locations submitted as parallel arrays
            locations_data = []
            warehouse_ids = form_data.getlist('location_warehouse_id[]')
            location_codes = form_data.getlist('location_code[]')
            safety_stocks = form_data.getlist('location_safety_stock[]')
            reorder_points = form_data.getlist('location_reorder_point[]')
            desired_days = form_data.getlist('location_desired_days_of_stock[]')
            moqs = form_data.getlist('location_moq[]')


            try:
                len_warehouses = len(warehouse_ids)
            except Exception:
                warehouse_ids = []
                location_codes = []
                safety_stocks = []
                reorder_points = []
                desired_days = []
                moqs = []


            for i in range(min(len(warehouse_ids), len(location_codes))):
                wid = warehouse_ids[i]
                code = location_codes[i].strip()
                # Extract and parse new inventory-specific parameters
                ss = int(safety_stocks[i]) if i < len(safety_stocks) and safety_stocks[i].isdigit() else 0
                rop = int(reorder_points[i]) if i < len(reorder_points) and reorder_points[i].isdigit() else 0
                dds = int(desired_days[i]) if i < len(desired_days) and desired_days[i].isdigit() else 30
                _moq = int(moqs[i]) if i < len(moqs) and moqs[i].isdigit() else 1


                if not wid or not code:
                    continue
                try:
                    wid_int = int(wid)
                except Exception:
                    continue
                locations_data.append({
                    'warehouse_id': wid_int,
                    'location_code': code,
                    'safety_stock': ss,
                    'reorder_point': rop,
                    'desired_days_of_stock': dds,
                    'moq': _moq
                })

            # Use Part.create to perform creation and location conflict checks
            result = Part.create(
                part_number=part_number,
                name=name,
                type=type_,
                description=description,
                unit=unit,
                quantity_per_box=quantity_per_box,
                locations_data=locations_data, # This now contains detailed inventory info
                lead_time=lead_time,
                standard_cost=standard_cost,
                # desired_days_of_stock and moq are no longer passed here
                is_active=True
            )

            # Part.create already returns a dict with success or error
            if not isinstance(result, dict):
                return {'success': False, 'error': '建立零件時發生錯誤', 'data': form_data}

            return result

        except Exception as e:
            db.session.rollback()
            print(f"Error in create_part_from_form: {e}")
            return {'success': False, 'error': f'建立零件時發生錯誤: {str(e)}', 'data': form_data}

    @staticmethod
    def update_part_from_form(part_id, form_data):
        """Updates a part from form data by delegating to Part.update."""
        try:
            # Extract all data from the form
            part_number = form_data.get('part_number', '').strip()
            name = form_data.get('name', '').strip()
            type_ = form_data.get('type', '').strip()
            description = form_data.get('description', '').strip()
            unit = form_data.get('unit', '').strip() or '個'
            quantity_per_box = int(form_data.get('quantity_per_box') or 1)
            lead_time = int(form_data.get('lead_time') or 0)
            standard_cost = float(form_data.get('standard_cost') or 0)
            # desired_days_of_stock and moq are now per location, not part of top-level Part creation

            # Basic validation
            if not part_number:
                return {'success': False, 'error': '零件編號為必填項目', 'data': form_data}
            if not name:
                return {'success': False, 'error': '零件名稱為必填項目', 'data': form_data}

            # Parse locations
            locations_data = []
            warehouse_ids = form_data.getlist('location_warehouse_id[]')
            location_codes = form_data.getlist('location_code[]')
            safety_stocks = form_data.getlist('location_safety_stock[]')
            reorder_points = form_data.getlist('location_reorder_point[]')
            desired_days = form_data.getlist('location_desired_days_of_stock[]')
            moqs = form_data.getlist('location_moq[]')


            for i in range(len(warehouse_ids)):
                wid = warehouse_ids[i]
                code = location_codes[i].strip()
                # Extract and parse new inventory-specific parameters
                ss = int(safety_stocks[i]) if i < len(safety_stocks) and safety_stocks[i].isdigit() else 0
                rop = int(reorder_points[i]) if i < len(reorder_points) and reorder_points[i].isdigit() else 0
                dds = int(desired_days[i]) if i < len(desired_days) and desired_days[i].isdigit() else 30
                _moq = int(moqs[i]) if i < len(moqs) and moqs[i].isdigit() else 1

                if not wid or not code:
                    continue
                try:
                    wid_int = int(wid)
                except (ValueError, TypeError):
                    continue
                locations_data.append({
                    'warehouse_id': wid_int,
                    'location_code': code,
                    'safety_stock': ss,
                    'reorder_point': rop,
                    'desired_days_of_stock': dds,
                    'moq': _moq
                })

            # Delegate to Part.update for all logic, including conflict checks
            result = Part.update(
                part_id=part_id,
                part_number=part_number,
                name=name,
                type=type_,
                description=description,
                unit=unit,
                quantity_per_box=quantity_per_box,
                locations_data=locations_data, # This now contains detailed inventory info
                lead_time=lead_time,
                standard_cost=standard_cost,
                # desired_days_of_stock and moq are no longer passed here
                is_active=True
            )

            # If the update fails, return the original form data for re-rendering
            if not result.get('success'):
                result['data'] = form_data
            
            return result

        except Exception as e:
            db.session.rollback()
            print(f"Error in update_part_from_form: {e}")
            return {'success': False, 'error': f'更新零件時發生錯誤: {str(e)}', 'data': form_data}

    @staticmethod
    def add_warehouse_location(form_data):
        warehouse_id_str = form_data.get('warehouse_id', '')
        location_code = form_data.get('location_code', '')
        description = form_data.get('description', '')
        
        if not warehouse_id_str or not location_code:
            return {'success': False, 'message': '倉庫和位置代碼為必填項目'}
        
        try:
            warehouse_id_int = int(warehouse_id_str)
            
            existing = WarehouseLocation.query.filter_by(
                warehouse_id=warehouse_id_int,
                location_code=location_code
            ).first()
            
            if existing:
                return {'success': False, 'message': '該倉位已存在'}
            
            new_location = WarehouseLocation(
                warehouse_id_int,
                location_code,
                description=description
            )
            db.session.add(new_location)
            db.session.commit()
            
            return {'success': True, 'message': '倉位新增成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉位新增失敗: {str(e)}'}

    @staticmethod
    def edit_warehouse_location(location_id, form_data):
        location_code = form_data.get('location_code')
        description = form_data.get('description', '')
        
        if not location_code:
            return {'success': False, 'message': '位置代碼為必填項目'}
        
        try:
            location = WarehouseLocation.query.get(location_id)
            if not location:
                return {'success': False, 'message': '找不到該倉位'}
            
            existing = WarehouseLocation.query.filter(
                WarehouseLocation.warehouse_id == location.warehouse_id,
                WarehouseLocation.location_code == location_code,
                WarehouseLocation.id != location_id
            ).first()
            
            if existing:
                return {'success': False, 'message': '該倉位代碼已存在於此倉庫'}
            
            location.location_code = location_code
            location.description = description
            db.session.commit()
            
            return {'success': True, 'message': '倉位更新成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉位更新失敗: {str(e)}'}

    @staticmethod
    def delete_warehouse_location(location_id):
        try:
            location = WarehouseLocation.query.get(location_id)
            if not location:
                return {'success': False, 'message': '找不到該倉位'}
            
            parts_using_assoc = PartWarehouseLocation.query.filter_by(
                warehouse_location_id=location_id
            ).all()
            
            if parts_using_assoc:
                part_list = []
                for assoc in parts_using_assoc:
                    part = Part.query.get(assoc.part_id)
                    if part:
                        part_list.append(f"{part.part_number} - {part.name}")
                
                if len(part_list) <= 5:
                    parts_info = '、'.join(part_list)
                else:
                    parts_info = '、'.join(part_list[:5]) + f' 等 {len(part_list)} 個零件'
                
                return {'success': False, 'message': f'無法刪除：此倉位被以下零件使用中：{parts_info}'}
            
            db.session.delete(location)
            db.session.commit()
            
            return {'success': True, 'message': '倉位刪除成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉位刪除失敗: {str(e)}'}

    @staticmethod
    def get_full_part_details(part_number):
        import logging
        logging.basicConfig(level=logging.INFO)
        logging.info(f"Fetching details for part_number: {part_number}")

        try:
            part = Part.get_by_part_number(part_number)
            if part is None:
                logging.warning(f"Part with part_number {part_number} not found.")
                return {'success': False, 'error': '找不到零件'}
            logging.info(f"Found part: {part.id}")

            order_history = OrderRegistration.query.options(
                joinedload(OrderRegistration.warehouse_location).joinedload(WarehouseLocation.warehouse)
            ).filter_by(part_number=part_number).order_by(OrderRegistration.created_at.desc()).limit(10).all()
            logging.info(f"Found {len(order_history)} order history records (limited to 10).")

            inventories = CurrentInventory.query.filter_by(part_id=part.id).all()
            logging.info(f"Found {len(inventories)} inventory records.")

            # Serialize data with individual error handling
            try:
                part_info = part.to_dict(include_locations=True)
            except Exception as e:
                logging.error(f"Error serializing part_info for part {part.id}: {e}", exc_info=True)
                raise

            try:
                order_history_data = [order.to_dict() for order in order_history]
            except Exception as e:
                logging.error(f"Error serializing order_history for part {part.id}: {e}", exc_info=True)
                raise

            try:
                # 為每個庫存位置加入消耗分析和訂購建議
                inventory_data = []
                for inv in inventories:
                    inv_dict = inv.to_dict()
                    
                    # 消耗分析 (基於30天工作日)
                    try:
                        inv_dict['consumption_analysis'] = inv.get_consumption_analysis(days=30)
                    except Exception as e:
                        logging.warning(f"Error calculating consumption analysis for inventory {inv.id}: {e}")
                        inv_dict['consumption_analysis'] = None
                    
                    # 訂購建議 (自動使用零件的 lead_time)
                    try:
                        inv_dict['order_suggestion'] = inv.get_order_suggestion()
                    except Exception as e:
                        logging.warning(f"Error calculating order suggestion for inventory {inv.id}: {e}")
                        inv_dict['order_suggestion'] = None
                    
                    # 儲位專屬訂單歷史 (最近 5 筆)
                    try:
                        location_orders = OrderRegistration.query.filter_by(
                            part_number=part_number,
                            warehouse_location_id=inv.warehouse_location_id
                        ).order_by(OrderRegistration.created_at.desc()).limit(5).all()
                        inv_dict['recent_orders'] = [order.to_dict() for order in location_orders]
                    except Exception as e:
                        logging.warning(f"Error fetching location orders for inventory {inv.id}: {e}")
                        inv_dict['recent_orders'] = []
                    
                    inventory_data.append(inv_dict)
                
            except Exception as e:
                logging.error(f"Error serializing inventory_data for part {part.id}: {e}", exc_info=True)
                raise

            # 計算零件級別的整體摘要
            try:
                summary = PartService._calculate_part_summary(inventories)
            except Exception as e:
                logging.warning(f"Error calculating part summary: {e}")
                summary = None

            result = {
                'part_info': part_info,
                'order_history': order_history_data,
                'inventories': inventory_data,
                'summary': summary  # 新增總體摘要
            }
            
            logging.info(f"Successfully fetched all details for part {part.id}")
            return {'success': True, 'data': result}

        except Exception as e:
            logging.error(f"Unhandled exception in get_full_part_details for part_number {part_number}: {e}", exc_info=True)
            # Re-raise the exception to let Flask handle it and produce a 500 error
            raise
    
    @staticmethod
    def _calculate_part_summary(inventories):
        """
        計算零件整體消耗摘要
        
        Args:
            inventories: CurrentInventory 物件列表
            
        Returns:
            dict: 整體摘要數據
        """
        if not inventories:
            return {
                'total_stock': 0,
                'total_available': 0,
                'overall_status': 'unknown',
                'min_days_of_stock': 0,
                'total_suggested_order': 0
            }
        
        total_stock = sum(inv.quantity_on_hand for inv in inventories)
        total_available = sum(inv.available_quantity for inv in inventories)
        
        # 收集所有儲位的狀態和庫存天數
        statuses = []
        days_of_stocks = []
        total_suggested = 0
        
        for inv in inventories:
            try:
                analysis = inv.get_consumption_analysis(days=30)
                suggestion = inv.get_order_suggestion()
                
                statuses.append(analysis['stock_status'])
                days_of_stocks.append(analysis['days_of_stock'])
                total_suggested += suggestion['suggested_quantity']
            except:
                continue
        
        # 整體狀態：只要有一個 critical 就是 critical
        if 'critical' in statuses:
            overall_status = 'critical'
        elif 'warning' in statuses:
            overall_status = 'warning'
        elif statuses:
            overall_status = 'healthy'
        else:
            overall_status = 'unknown'
        
        # 最少庫存天數
        min_days = min(days_of_stocks) if days_of_stocks else 0
        
        return {
            'total_stock': total_stock,
            'total_available': total_available,
            'overall_status': overall_status,
            'min_days_of_stock': round(min_days, 1),
            'total_suggested_order': int(total_suggested),
            'location_count': len(inventories)
        }

    @staticmethod
    def export_parts_excel(search='', sort_by='part_number', sort_order='asc'):
        """匯出零件清單為 Excel 檔案"""
        query = Part.query
        if sort_by == 'storage_location':
            query = query.outerjoin(PartWarehouseLocation).outerjoin(WarehouseLocation).outerjoin(Warehouse)

        if search:
            query = query.filter(or_(
                Part.name.ilike(f'%{search}%'),
                Part.part_number.ilike(f'%{search}%')
            ))
        
        valid_columns = ['part_number', 'name', 'type', 'description', 'unit', 'quantity_per_box', 'lead_time', 'storage_location']
        if sort_by not in valid_columns:
            sort_by = 'part_number'
        
        if sort_by == 'storage_location':
            if sort_order.lower() == 'desc':
                query = query.order_by(db.desc(func.coalesce(Warehouse.name, '')), db.desc(func.coalesce(WarehouseLocation.location_code, '')))
            else:
                query = query.order_by(func.coalesce(Warehouse.name, ''), func.coalesce(WarehouseLocation.location_code, ''))
        else:
            if sort_order.lower() == 'desc':
                query = query.order_by(db.desc(getattr(Part, sort_by)))
            else:
                query = query.order_by(getattr(Part, sort_by))

        parts = query.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "零件清單"

        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        headers = ['零件編號', '名稱', '類型', '備註', '單位', '每盒數量', '採購前置期 (天)', '儲存位置']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border

        for row_num, part in enumerate(parts, 2):
            locations = ", ".join([f"{assoc.warehouse_location.warehouse.name}-{assoc.warehouse_location.location_code}" for assoc in part.location_associations]) if part.location_associations else "無"
            
            ws.cell(row=row_num, column=1, value=part.part_number).border = border
            ws.cell(row=row_num, column=2, value=part.name).border = border
            ws.cell(row=row_num, column=3, value=part.type).border = border
            ws.cell(row=row_num, column=4, value=part.description).border = border
            ws.cell(row=row_num, column=5, value=part.unit).border = border
            ws.cell(row=row_num, column=6, value=part.quantity_per_box).border = border
            ws.cell(row=row_num, column=7, value=part.lead_time).border = border
            ws.cell(row=row_num, column=8, value=locations).border = border

        for col_num, column_cells in enumerate(ws.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            max_length = len(str(ws.cell(row=1, column=col_num).value))
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        from datetime import datetime
        filename = f"零件清單_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"

        return output.getvalue(), filename

    @staticmethod
    def get_part_autocomplete_suggestions(query):
        """Provides autocomplete suggestions for part numbers and names."""
        if not query or len(query) < 1:
            return []

        parts = Part.query.filter(
            or_(
                Part.part_number.ilike(f'%{query}%'),
                Part.name.ilike(f'%{query}%')
            )
        ).limit(10).all()

        results = [{'part_number': part.part_number, 'name': part.name} for part in parts]
        return results

    @staticmethod
    def add_warehouse(form_data):
        code = form_data.get('code', '')
        name = form_data.get('name', '')
        description = form_data.get('description', '')
        
        if not code or not name:
            return {'success': False, 'message': '倉庫編號和名稱為必填項目'}
        
        try:
            existing = Warehouse.query.filter_by(code=code).first()
            if existing:
                return {'success': False, 'message': '倉庫編號已存在'}
            
            new_warehouse = Warehouse(
                code,
                name,
                description=description,
                is_active=True
            )
            db.session.add(new_warehouse)
            db.session.commit()
            
            return {'success': True, 'message': '倉庫新增成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉庫新增失敗: {str(e)}'}

    @staticmethod
    def edit_warehouse(warehouse_id, form_data):
        name = form_data.get('name')
        description = form_data.get('description', '')
        
        if not name:
            return {'success': False, 'message': '倉庫名稱為必填項目'}
        
        try:
            warehouse = Warehouse.query.get(warehouse_id)
            if not warehouse:
                return {'success': False, 'message': '找不到該倉庫'}
            
            warehouse.name = name
            warehouse.description = description
            db.session.commit()
            
            return {'success': True, 'message': '倉庫更新成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉庫更新失敗: {str(e)}'}

    @staticmethod
    def delete_warehouse(warehouse_id):
        try:
            warehouse = Warehouse.query.get(warehouse_id)
            if not warehouse:
                return {'success': False, 'message': '找不到該倉庫'}
            
            locations_count = WarehouseLocation.query.filter_by(
                warehouse_id=warehouse_id
            ).count()
            
            if locations_count > 0:
                return {'success': False, 'message': f'無法刪除：此倉庫有 {locations_count} 個倉位，請先刪除所有倉位'}
            
            db.session.delete(warehouse)
            db.session.commit()
            
            return {'success': True, 'message': '倉庫刪除成功'}
        except Exception as e:
            db.session.rollback()
            return {'success': False, 'message': f'倉庫刪除失敗: {str(e)}'}

    @staticmethod
    def import_parts_from_excel(file_stream):
        """批量匯入零件從 Excel 檔案
        
        Excel 格式要求（與匯出格式一致）：
        - 零件編號 (必填)
        - 零件名稱 (必填)
        - 類型 (必填) - ZROM、A-事購件、B-生產工具、C-生產耗材
        - 備註
        - 單位 (必填)
        - 每盒數量
        - 採購前置期 (天)
        - 儲存位置 - 格式：倉別代碼:位置代碼，逗號分隔（可為空）
        
        Returns:
            dict: {
                'success': bool,
                'message': str,
                'error': str (if failed),
                'errors': list of str (individual row errors),
                'created_count': int,
                'updated_count': int,
                'total_rows': int
            }
        """
        try:
            # 讀取 Excel 檔案
            df = pd.read_excel(file_stream, sheet_name=0)
            
            # 清理欄位名稱：移除星號和多餘空白
            df.columns = df.columns.str.replace('*', '').str.strip()
            
            # 檢查必要欄位（處理可能的欄位名稱變化）
            # 支援兩種欄位名稱格式以向後相容
            column_mapping = {
                '零件編號': ['零件編號', 'Part Number'],
                '零件名稱': ['零件名稱', '名稱', 'Name'],
                '類型': ['類型', 'Type'],
                '備註': ['備註', '描述', 'Description'],
                '單位': ['單位', 'Unit'],
                '每盒數量': ['每盒數量', 'Quantity Per Box'],
                '採購前置期': ['採購前置期 (天)', '採購前置期', 'Lead Time'],
                '儲存位置': ['儲存位置', 'Storage Location', '儲存位置(倉別代碼:位置代碼, 逗號分隔)']
            }
            
            # 找到實際使用的欄位名稱
            actual_columns = {}
            for key, possible_names in column_mapping.items():
                for name in possible_names:
                    if name in df.columns:
                        actual_columns[key] = name
                        break
            
            # 檢查必填欄位
            required_fields = ['零件編號', '零件名稱', '類型', '單位']
            missing_fields = [field for field in required_fields if field not in actual_columns]
            
            if missing_fields:
                return {
                    'success': False,
                    'error': f'Excel 檔案缺少必要欄位：{", ".join(missing_fields)}',
                    'errors': []
                }
            
            total_rows = len(df)
            created_count = 0
            updated_count = 0
            errors = []
            
            # 有效的類型選項（與系統一致）
            valid_types = ['ZROM', 'A', 'B', 'C']
            
            for index, row in df.iterrows():
                try:
                    row_num = index + 2  # Excel row (1-indexed, +1 for header)
                    
                    # 取得必填欄位
                    part_number = str(row.get(actual_columns['零件編號'], '')).strip()
                    name = str(row.get(actual_columns['零件名稱'], '')).strip()
                    part_type = str(row.get(actual_columns['類型'], '')).strip()
                    unit = str(row.get(actual_columns['單位'], '')).strip()
                    
                    # 驗證必填欄位
                    if not part_number or part_number == 'nan':
                        errors.append(f'第 {row_num} 行：零件編號為必填項目')
                        continue
                    
                    if not name or name == 'nan':
                        errors.append(f'第 {row_num} 行：零件名稱為必填項目')
                        continue
                    
                    if not part_type or part_type == 'nan':
                        errors.append(f'第 {row_num} 行：類型為必填項目')
                        continue
                    
                    # 驗證類型是否有效
                    if part_type not in valid_types:
                        errors.append(f'第 {row_num} 行：類型「{part_type}」無效，請選擇：{", ".join(valid_types)}')
                        continue
                    
                    if not unit or unit == 'nan':
                        errors.append(f'第 {row_num} 行：單位為必填項目')
                        continue
                    
                    # 取得選填欄位
                    description = ''
                    if '備註' in actual_columns:
                        description = str(row.get(actual_columns['備註'], '')).strip()
                        if description == 'nan':
                            description = ''
                    
                    # 每盒數量
                    try:
                        if '每盒數量' in actual_columns:
                            quantity_per_box = int(row.get(actual_columns['每盒數量'], 1))
                        else:
                            quantity_per_box = 1
                    except (ValueError, TypeError):
                        quantity_per_box = 1
                    
                    # 採購前置期
                    try:
                        if '採購前置期' in actual_columns:
                            lead_time = int(row.get(actual_columns['採購前置期'], 5))
                        else:
                            lead_time = 5
                    except (ValueError, TypeError):
                        lead_time = 5
                    
                    # 解析儲存位置（選填）
                    locations_data = []
                    if '儲存位置' in actual_columns:
                        location_str = str(row.get(actual_columns['儲存位置'], '')).strip()
                        
                        if location_str and location_str != 'nan' and location_str != '無':
                            # 分割多個位置 (逗號分隔)
                            location_pairs = [loc.strip() for loc in location_str.split(',')]
                            
                            for loc_pair in location_pairs:
                                if not loc_pair:
                                    continue
                                
                                # 支援兩種格式：倉庫名-位置代碼 或 倉別代碼:位置代碼
                                if ':' in loc_pair:
                                    # 格式：W001:A-01-01
                                    parts = loc_pair.split(':')
                                    if len(parts) != 2:
                                        errors.append(f'第 {row_num} 行：儲存位置格式錯誤 "{loc_pair}"')
                                        continue
                                    
                                    warehouse_code = parts[0].strip()
                                    location_code = parts[1].strip()
                                elif '-' in loc_pair:
                                    # 格式：倉庫名-位置代碼（匯出格式）
                                    parts = loc_pair.split('-', 1)
                                    warehouse_name = parts[0].strip()
                                    location_code = parts[1].strip()
                                    
                                    # 根據倉庫名稱查找倉庫代碼
                                    warehouse = Warehouse.query.filter_by(name=warehouse_name).first()
                                    if not warehouse:
                                        errors.append(f'第 {row_num} 行：找不到倉庫「{warehouse_name}」')
                                        continue
                                    warehouse_code = warehouse.code
                                else:
                                    errors.append(f'第 {row_num} 行：儲存位置格式錯誤 "{loc_pair}"，應為「倉別代碼:位置代碼」或「倉庫名-位置代碼」')
                                    continue
                                
                                # 查找倉庫
                                warehouse = Warehouse.query.filter_by(code=warehouse_code).first()
                                if not warehouse:
                                    errors.append(f'第 {row_num} 行：找不到倉庫代碼 "{warehouse_code}"，請先在系統中建立倉庫')
                                    continue
                                
                                # 查找或建立儲位
                                location = WarehouseLocation.query.filter_by(
                                    warehouse_id=warehouse.id,
                                    location_code=location_code
                                ).first()
                                
                                if not location:
                                    # 自動建立儲位
                                    location = WarehouseLocation(
                                        warehouse.id,
                                        location_code,
                                        description=f'由批量匯入自動建立'
                                    )
                                    db.session.add(location)
                                    db.session.flush()  # 取得 location.id
                                
                                locations_data.append({
                                    'warehouse_id': warehouse.id,
                                    'location_code': location_code
                                })
                    
                    # 檢查零件是否已存在
                    existing_part = Part.get_by_part_number(part_number)
                    
                    if existing_part:
                        # 更新現有零件
                        result = Part.update(
                            part_id=existing_part.id,
                            part_number=part_number,
                            name=name,
                            type=part_type,
                            description=description,
                            unit=unit,
                            quantity_per_box=quantity_per_box,
                            locations_data=locations_data,
                            lead_time=lead_time,
                            standard_cost=existing_part.standard_cost,  # 保留原有值
                            is_active=True
                        )
                        
                        if result.get('success'):
                            updated_count += 1
                        else:
                            errors.append(f'第 {row_num} 行 ({part_number})：{result.get("error", "更新失敗")}')
                    else:
                        # 建立新零件
                        result = Part.create(
                            part_number=part_number,
                            name=name,
                            type=part_type,
                            description=description,
                            unit=unit,
                            quantity_per_box=quantity_per_box,
                            locations_data=locations_data,
                            lead_time=lead_time,
                            standard_cost=0,  # 預設值
                            is_active=True
                        )
                        
                        if result.get('success'):
                            created_count += 1
                        else:
                            errors.append(f'第 {row_num} 行 ({part_number})：{result.get("error", "建立失敗")}')
                
                except Exception as e:
                    errors.append(f'第 {row_num} 行：處理時發生錯誤 - {str(e)}')
                    continue
            
            # 提交所有變更
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                return {
                    'success': False,
                    'error': f'資料庫提交失敗: {str(e)}',
                    'errors': errors
                }
            
            # 產生結果訊息
            success_msg = f'匯入完成：新增 {created_count} 筆，更新 {updated_count} 筆'
            if errors:
                success_msg += f'，{len(errors)} 筆失敗'
            
            return {
                'success': True,
                'message': success_msg,
                'errors': errors,
                'created_count': created_count,
                'updated_count': updated_count,
                'total_rows': total_rows
            }
            
        except Exception as e:
            db.session.rollback()
            return {
                'success': False,
                'error': f'讀取 Excel 檔案失敗: {str(e)}',
                'errors': []
            }